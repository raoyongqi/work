import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# 设置数据目录和文件路径
DATA_DIR = 'data'  # 更新为实际的数据目录
input_file = os.path.join(DATA_DIR, 'random_data.xlsx')
output_file = os.path.join(DATA_DIR, 'merged_data.xlsx')

# 读取 Excel 文件
df = pd.read_excel(input_file)

# 将 DataFrame 写入新的 Excel 文件
df.to_excel(output_file, index=False, engine='openpyxl')

# 加载保存的 Excel 文件以设置单元格对齐和合并
wb = load_workbook(output_file)
ws = wb.active

# 合并相邻且值相同的单元格
def merge_cells(ws, col, start_row, end_row):
    ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)
    cell = ws.cell(row=start_row, column=col)
    cell.alignment = Alignment(horizontal='center', vertical='center')

# 遍历前3列，对“活跃”行进行合并操作
for col in range(1, 4):  # 前三列
    merge_start = None

    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=4).value == '活跃':  # 状态列为第四列
            if merge_start is None:
                merge_start = row
            current_value = ws.cell(row=row, column=col).value
            next_value = ws.cell(row=row + 1, column=col).value if row < ws.max_row else None
            
            if current_value != next_value:
                if merge_start is not None and merge_start != row:
                    merge_cells(ws, col, merge_start, row)
                merge_start = None
        else:
            if merge_start is not None:
                merge_cells(ws, col, merge_start, row - 1)
                merge_start = None

    # 合并最后一组单元格
    if merge_start is not None and ws.cell(row=ws.max_row, column=4).value == '活跃':
        merge_cells(ws, col, merge_start, ws.max_row)

# 保存文件
wb.save(output_file)

print(f"处理完成并保存到 '{output_file}' 文件。")
